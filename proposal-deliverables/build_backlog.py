"""Build the WBGT CrewSafe SG Product Backlog workbook.
Mirrors the structure of 'Example- Prioritisation Strategy & Project Status Report.xlsx'
(Features / Business Process / Technologies / Project Status) and adds a
prioritised Product Backlog sheet derived from the project plan section 15.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/Users/zhongchengtiong/Documents/Obsidian Vault/AD_wbgt/proposal-deliverables/WBGT-CrewSafe-SG-Product-Backlog.xlsx"

# ---- palette ----
NAVY   = "1F3B57"
BLUE   = "2E6FB0"
ORANGE = "E8873A"
LIGHT  = "EAF1F8"
BAND   = "F5F8FC"
GREEN  = "3F8F5B"
GREY   = "6B7A88"

thin = Side(style="thin", color="C9D4DF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(cell, fill=BLUE, color="FFFFFF", size=11):
    cell.font = Font(bold=True, color=color, size=size, name="Calibri")
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = border

def cell_fmt(cell, bold=False, fill=None, color="212121", wrap=True, align="left"):
    cell.font = Font(bold=bold, color=color, name="Calibri", size=10)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = border

def title_row(ws, text, span, fill=NAVY):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(row=1, column=1, value=text)
    c.font = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
    c.fill = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

wb = openpyxl.Workbook()

# =====================================================================
# Sheet 1: Product Backlog (prioritised)  -- primary deliverable
# =====================================================================
ws = wb.active
ws.title = "Product Backlog"
title_row(ws, "WBGT CrewSafe SG  —  Prioritised Product Backlog", 7)
headers = ["ID", "Priority (MoSCoW)", "Feature / User Story", "Story Points",
           "Implementation Technologies", "Acceptance Summary", "Sprint"]
for j, h in enumerate(headers, 1):
    hdr(ws.cell(row=2, column=j, value=h))
ws.row_dimensions[2].height = 30

backlog = [
    ("US-01","Must","Authenticate and see only my assigned site",5,
     "Web (React, TypeScript); Mobile (React Native, Expo); Backend (Spring Boot, Spring Security, JWT); PostgreSQL",
     "RBAC and negative authorization tests pass","1"),
    ("US-02","Must","Create a shift with workers, tasks and intensity",8,
     "Web (React, MUI); Backend (Spring Boot, Spring Data JPA); PostgreSQL",
     "Shift validation and assignments persist","1"),
    ("US-03","Must","Worker views shift and submits readiness flags",5,
     "Mobile (React Native, Expo); Backend (Spring Boot); PostgreSQL",
     "Only assigned worker can submit; no medical free text","1"),
    ("US-04","Must","Ingest and store WBGT / weather with freshness",8,
     "Backend (Spring Boot scheduler); NEA data.gov.sg APIs; PostgreSQL",
     "Live and fixture modes work; duplicate ingestion safe","1"),
    ("US-05","Must","Supervisor sees current conditions and active shift context",5,
     "Web (React, Chart.js); Backend (Spring Boot)",
     "Web view uses shared backend and shows freshness","1"),
    ("US-06","Must","System forecasts 30 / 60-minute WBGT",8,
     "Python ML service (FastAPI, pandas, scikit-learn / XGBoost, time-series regression)",
     "Baselines measured; versioned prediction returned","2"),
    ("US-07","Must","Evaluate the correct deterministic policy actions",8,
     "Backend (Java 21 deterministic rules engine); PostgreSQL policy config",
     "Boundary and mixed-worker tests pass","2"),
    ("US-08","Must","Supervisor receives an explainable agent draft plan",8,
     "Agentic AI (tool-calling LLM via Spring Boot adapter, JSON-schema guarded tools)",
     "No unsupported actions; policy / model references shown","2"),
    ("US-09","Must","Supervisor approves, edits or rejects a plan",5,
     "Web (React); Backend (Spring Boot)",
     "Draft and final versions retained; authorization enforced","2"),
    ("US-10","Must","Worker receives and acknowledges an approved action",8,
     "Mobile (React Native); Backend (idempotency keys); PostgreSQL",
     "Only affected workers see it; idempotency works","2"),
    ("US-11","Must","Worker logs rest / hydration or raises a concern",5,
     "Mobile (React Native); Backend (Spring Boot)",
     "Status updates in supervisor view","2"),
    ("US-12","Must","Supervisor monitors pending, late and completed actions",5,
     "Web (React, Chart.js); Backend (Spring Boot)",
     "Live state and alert count match backend records","2"),
    ("US-13","Must","Manager views compliance and response-time charts",8,
     "Web (React, Chart.js); Backend aggregation queries",
     "Metrics verified against seeded expected values","3"),
    ("US-14","Must","Manager views forecast accuracy and fallback status",5,
     "Web (Chart.js); Python ML evaluation metrics; Backend",
     "Metrics match evaluation artifact","3"),
    ("US-15","Must","Manager exports the audit timeline",5,
     "Backend (CSV / PDF export); Web; PostgreSQL append-only audit",
     "Export contains trace IDs, actors, rules and decisions","3"),
    ("US-16","Must","Safe degraded behaviour during dependency failure",8,
     "Backend fallback adapters (NEA / ML / LLM); persistence forecast",
     "NEA, ML and LLM failure scenarios pass","3"),
    ("US-17","Should","In-app reminder for an unacknowledged approved action",3,
     "Backend (scheduler); Mobile (in-app notification)",
     "Reminder cannot alter instruction","3"),
    ("US-18","Could","Worker views an approved instruction in another language",5,
     "Agentic AI (human-approved LLM template); Mobile (React Native)",
     "Human-approved template only","Stretch"),
]
r = 3
for row in backlog:
    for j, val in enumerate(row, 1):
        band = BAND if (r % 2 == 0) else "FFFFFF"
        c = ws.cell(row=r, column=j, value=val)
        cell_fmt(c, fill=band, align=("center" if j in (1,4,7) else "left"),
                 bold=(j==1))
        if j == 2:
            col = {"Must":GREEN,"Should":ORANGE,"Could":GREY}[val]
            c.font = Font(bold=True, color=col, size=10, name="Calibri")
            c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 42
    r += 1
widths = [8, 15, 34, 9, 46, 40, 8]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A3"

# =====================================================================
# Sheet 2: Features (feature -> technologies)  matches example "Features"
# =====================================================================
ws2 = wb.create_sheet("Features")
title_row(ws2, "Features  —  Proposed Features & Implementation Technologies", 2)
for j, h in enumerate(["Feature (use case / user story)", "Additional Details (Implementation Technologies)"], 1):
    hdr(ws2.cell(row=2, column=j, value=h), fill=BLUE)
ws2.row_dimensions[2].height = 26
features = [
    ("Authenticate & site-scoped access","Web (React, TS) + Mobile (React Native) + Backend (Spring Boot, Spring Security, JWT, PostgreSQL)"),
    ("Configure shift, tasks & assignments","Web (React, MUI), Backend (Spring Boot, Spring Data JPA, PostgreSQL)"),
    ("Complete worker readiness check","Mobile (React Native, Expo), Backend (Spring Boot, PostgreSQL)"),
    ("Ingest WBGT & weather with freshness","Backend (Spring Boot scheduler), NEA data.gov.sg APIs, PostgreSQL"),
    ("View current & forecast risk (live conditions)","Web (React, Chart.js), Backend (Spring Boot)"),
    ("Forecast 30 / 60-minute WBGT","Python ML (FastAPI, pandas, scikit-learn / XGBoost, time-series)"),
    ("Evaluate deterministic heat policy","Backend (Java 21 rules engine, PostgreSQL policy config)"),
    ("Generate explainable agent draft plan","Agentic AI (tool-calling LLM via backend adapter, JSON-schema tools)"),
    ("Approve / edit / reject plan","Web (React), Backend (Spring Boot)"),
    ("Dispatch & acknowledge worker action","Mobile (React Native), Backend (idempotency, PostgreSQL)"),
    ("Log rest / hydration & raise safety concern","Mobile (React Native), Backend (Spring Boot)"),
    ("Monitor completion (live status board)","Web (React, Chart.js), Backend (Spring Boot)"),
    ("Compliance & response-time dashboard","Web (React, Chart.js), Backend aggregation"),
    ("Model-performance dashboard","Web (Chart.js), Python ML metrics, Backend"),
    ("Export audit timeline (CSV / PDF)","Backend (export service), Web, PostgreSQL append-only audit"),
    ("Graceful degraded mode","Backend fallback adapters (NEA / ML / LLM)"),
    ("[Stretch] Multilingual instruction template","Agentic AI (human-approved LLM template), Mobile"),
    ("[Cross-cutting] Cloud deployment","Azure Container Apps, Azure Static Web Apps, Azure PostgreSQL, Blob Storage"),
    ("[Cross-cutting] DevSecOps pipeline","GitHub Actions CI/CD, SAST + secret scan, dependency & container scanning"),
]
r = 3
for name, tech in features:
    band = BAND if r % 2 == 0 else "FFFFFF"
    cell_fmt(ws2.cell(row=r, column=1, value=name), bold=True, fill=band)
    cell_fmt(ws2.cell(row=r, column=2, value=tech), fill=band)
    ws2.row_dimensions[r].height = 30
    r += 1
ws2.column_dimensions["A"].width = 40
ws2.column_dimensions["B"].width = 72
ws2.freeze_panes = "A3"

# =====================================================================
# Sheet 3: Business Process (actor -> features)  matches example
# =====================================================================
ws3 = wb.create_sheet("Business Process")
title_row(ws3, "Business Process  —  Heat-Safety Operations Workflow", 2)
for j, h in enumerate(["Actor", "Features"], 1):
    hdr(ws3.cell(row=2, column=j, value=h), fill=BLUE)
ws3.row_dimensions[2].height = 24
bp = [
    ("SECTION","Core observe → decide → approve → dispatch → acknowledge → audit loop"),
    ("NEA Weather Service","Ingest WBGT & weather (system-triggered)"),
    ("System (ML)","Forecast 30 / 60-minute WBGT"),
    ("System (Policy)","Evaluate deterministic heat policy"),
    ("System (Agent)","Draft explainable intervention plan"),
    ("Site Supervisor","Review & approve / edit / reject plan"),
    ("Site Supervisor","Configure shift & tasks"),
    ("Site Supervisor","Monitor current & forecast risk"),
    ("Site Supervisor","Monitor completion & send approved reminder"),
    ("Outdoor Worker","Complete readiness check"),
    ("Outdoor Worker","Receive & acknowledge action"),
    ("Outdoor Worker","Log rest / hydration"),
    ("Outdoor Worker","Raise safety concern"),
    ("SECTION","Oversight & governance"),
    ("Safety Manager","Configure policy & users"),
    ("Safety Manager","View dashboards & audit / export timeline"),
]
r = 3
for actor, feat in bp:
    if actor == "SECTION":
        ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        c = ws3.cell(row=r, column=1, value=feat)
        cell_fmt(c, bold=True, fill=LIGHT, color=NAVY)
    else:
        band = BAND if r % 2 == 0 else "FFFFFF"
        cell_fmt(ws3.cell(row=r, column=1, value=actor), bold=True, fill=band)
        cell_fmt(ws3.cell(row=r, column=2, value=feat), fill=band)
    ws3.row_dimensions[r].height = 22
    r += 1
ws3.column_dimensions["A"].width = 24
ws3.column_dimensions["B"].width = 66
ws3.freeze_panes = "A3"

# =====================================================================
# Sheet 4: Technologies (architecture-centric)
# =====================================================================
ws4 = wb.create_sheet("Technologies")
title_row(ws4, "Architecture-Centric Features  —  Technology Stack", 3)
for j, h in enumerate(["Layer", "Technology", "Purpose in WBGT CrewSafe SG"], 1):
    hdr(ws4.cell(row=2, column=j, value=h), fill=BLUE)
ws4.row_dimensions[2].height = 24
tech = [
    ("Web app","React, TypeScript, Vite, Material UI, Chart.js","Supervisor / manager / admin console, dashboards, approval workflow"),
    ("Mobile app","React Native, Expo, TypeScript","Worker readiness, instructions, acknowledgement, rest / hydration logs"),
    ("Shared backend","Java 21, Spring Boot, Spring Security, Spring Data JPA","System of record, RBAC, policy engine host, audit"),
    ("Machine learning","Python, FastAPI, pandas, scikit-learn, XGBoost","30 / 60-minute WBGT time-series forecasting service (internal)"),
    ("Agentic AI","Tool-calling LLM via backend adapter, JSON schemas","Draft prioritised plans from guarded, authorised tools only"),
    ("Database","PostgreSQL","Users, sites, shifts, observations, recommendations, append-only audit"),
    ("API contract","OpenAPI 3, REST /api/v1","Contract between clients and shared backend"),
    ("Cloud","Azure Static Web Apps, Azure Container Apps, Azure PostgreSQL, Blob Storage","Hosting for web, backend, internal ML service and database"),
    ("Observability","Structured JSON logs, correlation IDs, Application Insights","Health checks, traceability, integration-failure metrics"),
    ("DevSecOps","GitHub Actions, SAST, secret / dependency / container scanning","CI/CD with automated tests and security remediation evidence"),
]
r = 3
for layer, t, purpose in tech:
    band = BAND if r % 2 == 0 else "FFFFFF"
    cell_fmt(ws4.cell(row=r, column=1, value=layer), bold=True, fill=band)
    cell_fmt(ws4.cell(row=r, column=2, value=t), fill=band)
    cell_fmt(ws4.cell(row=r, column=3, value=purpose), fill=band)
    ws4.row_dimensions[r].height = 30
    r += 1
ws4.column_dimensions["A"].width = 18
ws4.column_dimensions["B"].width = 46
ws4.column_dimensions["C"].width = 56
ws4.freeze_panes = "A3"

# =====================================================================
# Sheet 5: Project Status (tracking template)
# =====================================================================
ws5 = wb.create_sheet("Project Status")
title_row(ws5, "Project Status Report  (C = Completed · I = In Progress · blank = Not started)", 7)
cols = ["Feature","Implementation Technologies","Sprint","Requirements","Development","Integration Testing","Deployed to Cloud"]
for j, h in enumerate(cols, 1):
    hdr(ws5.cell(row=2, column=j, value=h), fill=BLUE)
ws5.row_dimensions[2].height = 30
status = [
    ("Authenticate & site-scoped access","Web + Mobile + Backend, PostgreSQL","1","C","C","I",""),
    ("Configure shift, tasks & assignments","Web, Backend, PostgreSQL","1","C","I","",""),
    ("Complete worker readiness check","Mobile, Backend, PostgreSQL","1","C","I","",""),
    ("Ingest WBGT & weather with freshness","Backend, NEA APIs, PostgreSQL","1","C","C","",""),
    ("View current & forecast risk","Web (Chart.js), Backend","1","C","","",""),
    ("Forecast 30 / 60-minute WBGT","Python ML (FastAPI, XGBoost)","2","C","","",""),
    ("Evaluate deterministic heat policy","Backend rules engine","2","C","","",""),
    ("Generate explainable agent draft plan","Agentic AI (LLM adapter)","2","","","",""),
    ("Approve / edit / reject plan","Web, Backend","2","","","",""),
    ("Dispatch & acknowledge worker action","Mobile, Backend","2","","","",""),
    ("Log rest / hydration & raise concern","Mobile, Backend","2","","","",""),
    ("Monitor completion (live status)","Web, Backend","2","","","",""),
    ("Compliance & response-time dashboard","Web (Chart.js), Backend","3","","","",""),
    ("Model-performance dashboard","Web, Python ML metrics","3","","","",""),
    ("Export audit timeline","Backend (CSV/PDF), Web","3","","","",""),
    ("Graceful degraded mode","Backend fallback adapters","3","","","",""),
]
r = 3
for row in status:
    band = BAND if r % 2 == 0 else "FFFFFF"
    for j, val in enumerate(row, 1):
        c = ws5.cell(row=r, column=j, value=val)
        align = "center" if j >= 3 else "left"
        cell_fmt(c, bold=(j==1), fill=band, align=align)
        if j >= 4 and val == "C":
            c.font = Font(bold=True, color=GREEN, size=10)
        elif j >= 4 and val == "I":
            c.font = Font(bold=True, color=ORANGE, size=10)
    ws5.row_dimensions[r].height = 26
    r += 1
w5 = [34, 34, 8, 13, 13, 16, 16]
for i, w in enumerate(w5, 1):
    ws5.column_dimensions[get_column_letter(i)].width = w
ws5.freeze_panes = "C3"

wb.save(OUT)
print("Saved:", OUT)
print("Sheets:", wb.sheetnames)
